import tkinter
from tkinter import messagebox
import openpyxl
import hashlib

window = tkinter.Tk()
window.title("Sign in")
window.geometry('340x440')
window.configure(bg='#FFA500')

def login():
    # Open the Excel file
    filepath = "..\\DB.xlsx"
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
    except FileNotFoundError:
        messagebox.showerror(title="Error", message="Database file not found.")
        return

    # Get entered username (or email) and password
    entered_username = username_entry.get()
    entered_password = hashlib.sha256(password_entry.get().encode()).hexdigest()

    # Search for the username (or email) and password in the Excel file
    for row in sheet.iter_rows(min_row=2, max_col=3, max_row=sheet.max_row, values_only=True):
        if row[0] == entered_username or row[2] == entered_username:
            if row[1] == entered_password:
                messagebox.showinfo(title="Login Success", message="You successfully logged in.")
                return
    messagebox.showerror(title="Error", message="Invalid login.")

frame = tkinter.Frame(bg='#FFA500')

# Creating widgets
login_label = tkinter.Label(frame, text="Login", bg='#FFA500', fg="#FF3399", font=("Arial", 30))
username_label = tkinter.Label(frame, text="Username or Email", bg='#FFA500', fg="#FFFFFF", font=("Arial", 16))
username_entry = tkinter.Entry(frame, font=("Arial", 16))
password_entry = tkinter.Entry(frame, show="*", font=("Arial", 16))
password_label = tkinter.Label(frame, text="Password", bg='#FFA500', fg="#FFFFFF", font=("Arial", 16))
login_button = tkinter.Button(frame, text="Login", bg="#FF3399", fg="#FFFFFF", font=("Arial", 14), command=login)

# Placing widgets on the screen
login_label.grid(row=0, column=0, columnspan=2, sticky="news", pady=40)
username_label.grid(row=1, column=0)
username_entry.grid(row=1, column=1, pady=20)
password_label.grid(row=2, column=0)
password_entry.grid(row=2, column=1, pady=20)
login_button.grid(row=3, column=0, columnspan=2, pady=30)

frame.pack()

window.mainloop()

